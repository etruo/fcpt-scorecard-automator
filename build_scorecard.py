### Packages: 

import pdfplumber
import re
from openai import OpenAI
import json
import re
from openpyxl import load_workbook
from dotenv import load_dotenv
from typing import Optional
from pathlib import Path
import textwrap
from extractor import get_best_payload, extract_plain_text
import os
import tempfile
from datetime import datetime, date, timedelta
from extractor import get_best_payload, extract_plain_text
import boto3
from botocore.exceptions import ClientError
import io

load_dotenv()
#openai_key = os.environ["OPENAI_API_KEY"]  # Works locally 
#print(os.getenv("OPENAI_API_KEY")[:11])  # Just print a few characters for safety

REQUIRED_KEYS = {
    "Lease Structure",
    "Lease Term",
    "Absolute Rent",
    "Rent Growth",
    "Acreage",
    "Restaurant/Auto/Medical?",
    "Single Tenant?",
    "Drive-Thru (QSR) / Carry-out (CDR)",
    "Box Size",
    "Address",
    "Year Built",
    "Current Tenant",
    "Number of National Locations",
}

KW = [
    "LEASE", "RENT", "ACRE", "ADDRESS", "TENANT",
    "CURRENT", "GLA", "CAP RATE", "YEAR BUILT", "DRIVE-THRU", "CARRY-OUT"
]

def extract_address(text):
    """
    Extracts the property address from the OM text.
    Args:
        text (str): Full text of the OM.
    Returns:
        str: Extracted address.
    """
    import re
    # Regex pattern to identify the address line
    match = re.search(r"(Address|Property Address):\s*(.+)", text, re.IGNORECASE)
    return match.group(2).strip() if match else None

def calculate_remaining_term(expiration_date_str: str) -> float:
    """
    Calculate the remaining lease term in years from today to the expiration date.
    
    Args:
        expiration_date_str: Date string in various formats (e.g., "June 2029", "6/2029", "06/30/2029")
    
    Returns:
        float: Number of years remaining (rounded to 2 decimal places)
    """
    try:
        # Try to parse various date formats
        expiration_date = None
        
        # Try to match "Month Year" format (e.g., "June 2029")
        month_year_match = re.match(r"(\w+)\s+(\d{4})", expiration_date_str)
        if month_year_match:
            month_str, year_str = month_year_match.groups()
            # Convert month name to number (assuming it's a valid month name)
            month_map = {
                'january': 1, 'february': 2, 'march': 3, 'april': 4,
                'may': 5, 'june': 6, 'july': 7, 'august': 8,
                'september': 9, 'october': 10, 'november': 11, 'december': 12
            }
            month = month_map[month_str.lower()]
            expiration_date = date(int(year_str), month, 1)
            # Set to end of month since specific day isn't provided
            if month < 12:
                next_month = date(int(year_str), month + 1, 1)
            else:
                next_month = date(int(year_str) + 1, 1, 1)
            expiration_date = next_month - timedelta(days=1)
        
        # Try to match "MM/YYYY" format
        elif re.match(r"\d{1,2}/\d{4}$", expiration_date_str):
            month, year = map(int, expiration_date_str.split('/'))
            expiration_date = date(year, month, 1)
            # Set to end of month
            if month < 12:
                next_month = date(year, month + 1, 1)
            else:
                next_month = date(year + 1, 1, 1)
            expiration_date = next_month - timedelta(days=1)
        
        # Try to match "MM/DD/YYYY" format
        elif re.match(r"\d{1,2}/\d{1,2}/\d{4}$", expiration_date_str):
            month, day, year = map(int, expiration_date_str.split('/'))
            expiration_date = date(year, month, day)
        
        if expiration_date is None:
            raise ValueError(f"Could not parse date: {expiration_date_str}")
        
        # Calculate years between today and expiration
        today = date.today()
        days_remaining = (expiration_date - today).days
        years_remaining = days_remaining / 365.25  # Using 365.25 to account for leap years
        
        return round(years_remaining, 2)
    
    except Exception as e:
        print(f"Error calculating lease term: {e}")
        return None

def interpret_payload_with_gpt(payload: str, *, client: OpenAI):
    """
    Uses the new OpenAI client interface to extract key fields from the provided best_tables.
    
    Args:
        best_tables (list): A list of dictionaries containing extracted table data.
        model (str): The model to use (default "gpt-4o").
    
    Returns:
        dict: A dictionary mapping the following keys to their extracted values.
              If a field is not found, its value will be null.
    """
    model ="gpt-4o"
    required = REQUIRED_KEYS
    
    prompt = textwrap.dedent(f"""
        You are an expert data extractor. 

        Please extract and interpret the following key fields:

        1. Lease Structure  
        2. Lease Term (IMPORTANT: Extract the exact lease expiration date AND calculate remaining years)  
        3. Absolute Rent  
        4. Rent Growth  
        5. Acreage  
        6. Is Current Tenant a Restaurant, Auto, or Medical Facility? (Yes or No)  
        7. Is The Building Currently a Single Tenant Building? (Yes or No)  
        8. Does the operator have a Drive-Thru (QSR) or Carry-out (CDR) available?  
        9. Box Size  
        10. Address (Split into dictionary with the following keys: 'Line 1', 'City', 'State', 'Zip')  
        11. Year Built  
        12. Current Tenant (restaurant, auto shop, medical operator name)  
        13. Number of National Locations  

        Instructions:
        - If a field is not found in the tables, set its value to null.  
        - For Lease Term: You MUST extract the exact lease expiration date (e.g., "June 2029", "6/30/2029").
        - For Lease Term: The value should be a dictionary with two keys:
          * "expiration_date": The exact expiration date as found in the text
          * "remaining_years": Calculate the years between today and the expiration date
        - For field 4: List the percent growth. If it is only one bump over a multi year period, calculate the average annual bump (ex. 10% over five years = 2%)
        - For field 4: If the rent growth is not listed, calculate the in-place percentage growth of the current rent rate. 
        - For field 7: if the answer is No, also note whether the building is >50%, >33%, or <33% restaurant.  
        - For field 8: if the tenant is a restaurant, determine if it qualifies as QSR or CDR. If not a restaurant, return "NA".  
        - For field 12: prefer the franchise name over the operator's legal entity name.  
        - For field 13: estimate the national presence and provide the number of U.S. locations of the specific type of restaurant, auto shop, or medical clinic.

        Return only a JSON object that exactly matches the schema below. If a value is unknown, return null.

        Schema (example format):

        ```json
        {{
        "Lease Structure": "NNN",
        "Lease Term": {{
            "expiration_date": "June 2029",
            "remaining_years": 5.4
        }},
        "Absolute Rent": 120000,
        "Rent Growth": "3% annually",
        "Acreage": 0.83,
        "Restaurant/Auto/Medical?": "Yes",
        "Single Tenant?": "Yes",
        "Drive-Thru (QSR) / Carry-out (CDR)": "QSR",
        "Box Size": "2,300 sqft",
        "Address": {{
            "Line 1": "123 Main St",
            "City": "Cedar Rapids",
            "State": "IA",
            "Zip": "52404"
        }},
        "Year Built": 2015,
        "Current Tenant": "Taco Bell",
        "Number of National Locations": 7500
        }} 
        --DATA START--
        {payload}
        --DATA END--
    """)

    # Create the chat completion using the new client interface.
    resp = client.chat.completions.create(
        model=model,
        messages=[{"role": "user", "content": prompt}],
        response_format={"type":"json_object"},
        temperature=0
    )
    try:
        data = json.loads(resp.choices[0].message.content)
        
        # Post-process the lease term if it exists
        if isinstance(data.get("Lease Term"), dict):
            expiration_date = data["Lease Term"].get("expiration_date")
            if expiration_date:
                # Calculate the remaining years using our helper function
                remaining_years = calculate_remaining_term(expiration_date)
                if remaining_years is not None:
                    data["Lease Term"]["remaining_years"] = remaining_years
                    # Also store the raw years for the scoring function
                    data["Lease Term"] = str(remaining_years)
    except Exception as e:
        print(f"Error processing GPT response: {e}")
        data = {}

    # guarantee all keys exist
    safe = {k: data.get(k) for k in required}
    if safe["Address"] is None:
        safe["Address"] = {"Line 1":None,"City":None,"State":None,"Zip":None}
    return safe

def normalize_fields(fields):
    # Ensure every field exists, and numeric-looking strings
    # get turned into strings, so your downstream code can
    # use parse_numeric_value uniformly.
    defaults = {
        "Absolute Rent": "",
        "Rent Growth": "",
        "Acreage": "",
        "Box Size": "",
        "Number of National Locations": "",
    }
    out = {}
    for k, default in defaults.items():
        v = fields.get(k, default)
        # If the model gave you a raw number, turn it back into a string
        if isinstance(v, (int, float)):
            v = str(v)
        out[k] = v

    # Copy over everything else untouched
    for k in fields:
        if k not in out:
            out[k] = fields[k]
    return out


###########################################
# Helper Functions for Parsing & Mapping  #
###########################################

def parse_numeric_value(value):
    """
    Now handles:
      • None or ""      -> 0
      • int or float    -> returned as is
      • "1,234.56 SF"   -> 1234.56
    """
    if value is None:
        return 0

    # ←——  NEW: if they gave you a number, just use it
    if isinstance(value, (int, float)):
        return value

    s = str(value).replace(',', '')
    match = re.search(r"[\d\.]+", s)
    if match:
        num_str = match.group(0)
        return float(num_str) if '.' in num_str else int(num_str)
    return 0

def map_address(addr_dict: dict, tenant: str) -> str:
    """
    Returns the string for cell C3:

        FCPT Scorecard: {tenant}, {Line 1} {City}, {State} {Zip}
    """
    if not isinstance(addr_dict, dict):
        return f"FCPT Scorecard: {tenant}"          # fallback

    concat = (
        f"{addr_dict.get('Line 1','')} "
        f"{addr_dict.get('City','')}, "
        f"{addr_dict.get('State','')} "
        f"{addr_dict.get('Zip','')}"
    ).strip()
    return f"FCPT Scorecard: {tenant}, {concat}"


def map_restaurant_auto_medical(val):
    """
    Maps the Restaurant/Auto/Medical? field.
    If value is 'Yes', return integer 2. Otherwise, return 0.
    """
    if val and val.strip().lower() == "yes":
        return 2
    return 0

def map_single_tenant(val):
    """
    Maps the Single Tenant? field.
    If value is 'Yes', return integer 2.
    If 'No', you might need more input; here we default to 0.
    """
    if val and val.strip().lower() == "yes":
        return 2
    # For "No", further breakdown might be applied.
    # For now, default to 0.
    return 0

def map_portfolio_target():
    """Defaults for both Portfolio Management Target Brand and Geography."""
    return 1

def map_acreage(acreage_str):
    """
    Maps Acreage to the appropriate score based on bounds:
      >2.25   => 7
      1.75 - 2.25 => 6
      1.25 - 1.75 => 5
      0.75 - 1.25 => 4
      0.5  - 0.75 => 2
      <0.5   => 0
    """
    acreage = parse_numeric_value(acreage_str)
    if acreage > 2.25:
        return 7
    elif 1.75 <= acreage <= 2.25:
        return 6
    elif 1.25 <= acreage < 1.75:
        return 5
    elif 0.75 <= acreage < 1.25:
        return 4
    elif 0.5 <= acreage < 0.75:
        return 2
    else:
        return 0

def map_drive_thru_carryout(val):
    """
    Maps the Drive-Thru (QSR) / Carry-out (CDR) field.
    Here we assume if the value indicates QSR or contains "yes", return 2.
    If it indicates CDR or "no", return 0.
    (You may adjust this logic based on your data.)
    """
    if val and ("qsr" in val.lower() or val.strip().lower() == "NA" or "cdr" in val.lower()):
        return 2
    return 0

def map_box_size(box_size_str):
    """
    Maps Box Size (in square feet) to:
      0 if box size is less than 2000 sqft or greater than 9000 sqft,
      otherwise, return 2.
    """
    sqft = parse_numeric_value(box_size_str)
    if sqft < 2000 or sqft > 9000:
        return 0
    else:
        return 2

def map_national_locations(extracted_fields):
    """
    Accepts either:
      - a dict under "Concept Durability" containing the key
      - or a top‑level "Number of National Locations" string
    """
    # Try nested first
    if isinstance(extracted_fields, dict) and "Number of National Locations" in extracted_fields:
        loc_str = extracted_fields["Number of National Locations"]
    else:
        # Fallback: maybe they passed the whole `extracted_fields`
        loc_str = extracted_fields.get("Number of National Locations", "")

    locations = parse_numeric_value(loc_str)
    print(f"Gathered number of locations: {locations}")

    if locations > 600:
        return 2.5
    elif 300 < locations <= 600:
        return 1.5
    elif 100 < locations <= 300:
        return 1
    else:
        return 0
        
def map_lease_structure(val):
    """   
    Maps Lease Structure using the table:
      Master Lease -> 1.5
      NNN          -> 1
      NN           -> 0
      Meaningful LL Obligations -> (set to 0, or adjust if needed)
    """
    if not val:
        return 0
    val = val.strip().lower()
    if "master lease" in val:
        return 1.5
    elif "nnn" in val:
        return 1
    elif val == "nn":
        return 0
    elif "meaningful ll obligations" in val:
        return 0  # or assign a different value if desired
    return 0

def map_lease_term(term_str):
    """
    Maps Lease Term in years to a score:
      >15 years => 5
      10-15     => 4
      7.5-10    => 3
      5-7.5     => 2
      3-5       => 1
      <3        => 0
    """
    years = parse_numeric_value(term_str)
    if years > 15:
        return 5
    elif 10 <= years <= 15:
        return 4
    elif 7.5 <= years < 10:
        return 3
    elif 5 <= years < 7.5:
        return 2
    elif 3 <= years < 5:
        return 1
    else:
        return 0

def map_absolute_rent(rent_str, building_type):
    """
    Maps Absolute Rent based on the building type.
    For instance, if building_type == "CDR" (carry-out), use:
      - < $170k -> 8
      - $170k to $210k -> 7
      - $210k to $250k -> 6
      - $250k to $285k -> 5
      - $285k to $330k -> 3
      - > $330k -> 0
    (You can add other tables for QSR, FCR, etc.)
    """
    rent = parse_numeric_value(rent_str)
    if building_type == "CDR":
        if rent < 170000:
            return 8
        elif 170000 <= rent < 210000:
            return 7
        elif 210000 <= rent < 250000:
            return 6
        elif 250000 <= rent < 285000:
            return 5
        elif 285000 <= rent < 330000:
            return 3
        else:
            return 0

    if building_type == "QSR":
        if rent < 90000:
            return 8
        elif 90000 <= rent < 110000:
            return 7
        elif 110000 <= rent < 135000:
            return 6
        elif 135000 <= rent < 150000:
            return 5
        elif 150000 <= rent < 170000:
            return 3
        else:
            return 0
    # Default if building_type not recognized:
    return 0

def map_rent_growth(rent_growth_str):
    """
    Parse the rent growth percentage value.
    For example, "1.5% Annually" -> 1.5.
    You might later choose to map this into a score if desired.
    """
    rent_growth = parse_numeric_value(rent_growth_str)

    if rent_growth < 0.5 or rent_growth > 2.25:
        return 0
    elif 0.5 <= rent_growth < 1.25:
        return 1
    elif 1.25 <= rent_growth <= 2.25:
        return 1.5
    else:
        return 0 

def map_restaurant_auto_medical_comment(tenant: str, is_ram: str) -> str:
    """Maps the comment for Restaurant/Auto/Medical field"""
    if is_ram and is_ram.strip().lower() == "yes" and tenant:
        return f"{tenant} is a restaurant/auto/medical tenant"
    return "Not a restaurant/auto/medical tenant"

def map_single_tenant_comment(val: str) -> str:
    """Maps the comment for Single Tenant field"""
    if val and val.strip().lower() == "yes":
        return "Free-standing, single-tenant asset"
    return "Multi-tenant property"

def map_portfolio_target_brand_comment(tenant: str) -> str:
    """Maps the comment for Portfolio Target Brand"""
    return f"{tenant} is a target brand for FCPT"

def map_portfolio_target_geography_comment(address: dict) -> str:
    """Maps the comment for Portfolio Target Geography"""
    state = address.get("State", "")
    if state:
        return f"{state} is an attractive market for FCPT"
    return "Location is an attractive market for FCPT"

def map_acreage_comment(acreage: str) -> str:
    """Maps the comment for Acreage"""
    acres = parse_numeric_value(acreage)
    if acres > 2.25:
        return f"Large parcel size of {acres} acres"
    elif 1.75 <= acres <= 2.25:
        return f"Good parcel size of {acres} acres"
    elif 1.25 <= acres < 1.75:
        return f"Adequate parcel size of {acres} acres"
    elif 0.75 <= acres < 1.25:
        return f"Moderate parcel size of {acres} acres"
    elif 0.5 <= acres < 0.75:
        return f"Small parcel size of {acres} acres"
    else:
        return f"Very small parcel size of {acres} acres"

def map_drive_thru_comment(val: str, tenant: str) -> str:
    """Maps the comment for Drive-Thru/Carry-out"""
    if not val or val.lower() == "na":
        return "Not applicable for this tenant type"
    val_lower = val.lower()
    if "qsr" in val_lower:
        return f"{tenant} has a drive-thru"
    elif "cdr" in val_lower:
        return f"{tenant} has carry-out capability"
    return "No drive-thru or carry-out capability"

def map_box_size_comment(size: str) -> str:
    """Maps the comment for Box Size"""
    sqft = parse_numeric_value(size)
    if 2000 <= sqft <= 9000:
        return f"Optimal box size of {sqft:,.0f} square feet"
    elif sqft < 2000:
        return f"Box size of {sqft:,.0f} square feet is below optimal range"
    else:
        return f"Box size of {sqft:,.0f} square feet is above optimal range"

def map_national_locations_comment(locations: str, tenant: str) -> str:
    """Maps the comment for Number of National Locations"""
    loc_num = parse_numeric_value(locations)
    if loc_num > 600:
        return f"{tenant} has strong national presence with {loc_num:,.0f} locations"
    elif 300 < loc_num <= 600:
        return f"{tenant} has moderate national presence with {loc_num:,.0f} locations"
    elif 100 < loc_num <= 300:
        return f"{tenant} has limited national presence with {loc_num:,.0f} locations"
    else:
        return f"{tenant} has minimal national presence with {loc_num:,.0f} locations"

def map_lease_structure_comment(structure: str) -> str:
    """Maps the comment for Lease Structure"""
    if not structure:
        return "Lease structure not specified"
    structure_lower = structure.strip().lower()
    if "master lease" in structure_lower:
        return "Master lease provides additional tenant credit support"
    elif "nnn" in structure_lower:
        return "NNN lease structure"
    elif structure_lower == "nn":
        return "NN lease structure with some landlord responsibilities"
    elif "meaningful ll obligations" in structure_lower:
        return "Significant landlord obligations under lease"
    return "Standard lease structure"

def map_lease_term_comment(term: str) -> str:
    """Maps the comment for Lease Term"""
    years = parse_numeric_value(term)
    if years > 15:
        return f"Very long remaining lease term of {years:,.1f} years"
    elif 10 <= years <= 15:
        return f"Long remaining lease term of {years:,.1f} years"
    elif 7.5 <= years < 10:
        return f"Moderate remaining lease term of {years:,.1f} years"
    elif 5 <= years < 7.5:
        return f"Short remaining lease term of {years:,.1f} years"
    elif 3 <= years < 5:
        return f"Very short remaining lease term of {years:,.1f} years"
    else:
        return f"Extremely short remaining lease term of {years:,.1f} years"

def map_absolute_rent_comment(rent: str, building_type: str) -> str:
    """Maps the comment for Absolute Rent"""
    rent_val = parse_numeric_value(rent)
    if building_type == "CDR":
        if rent_val < 170000:
            return f"Very attractive rent of ${rent_val:,.0f} for CDR"
        elif 170000 <= rent_val < 210000:
            return f"Attractive rent of ${rent_val:,.0f} for CDR"
        elif 210000 <= rent_val < 250000:
            return f"Moderate rent of ${rent_val:,.0f} for CDR"
        elif 250000 <= rent_val < 285000:
            return f"High rent of ${rent_val:,.0f} for CDR"
        elif 285000 <= rent_val < 330000:
            return f"Very high rent of ${rent_val:,.0f} for CDR"
        else:
            return f"Extremely high rent of ${rent_val:,.0f} for CDR"
    elif building_type == "QSR":
        if rent_val < 90000:
            return f"Very attractive rent of ${rent_val:,.0f} for QSR"
        elif 90000 <= rent_val < 110000:
            return f"Attractive rent of ${rent_val:,.0f} for QSR"
        elif 110000 <= rent_val < 135000:
            return f"Moderate rent of ${rent_val:,.0f} for QSR"
        elif 135000 <= rent_val < 150000:
            return f"High rent of ${rent_val:,.0f} for QSR"
        elif 150000 <= rent_val < 170000:
            return f"Very high rent of ${rent_val:,.0f} for QSR"
        else:
            return f"Extremely high rent of ${rent_val:,.0f} for QSR"
    return f"Rent: ${rent_val:,.0f}"

def map_rent_growth_comment(growth: str) -> str:
    """Maps the comment for Rent Growth"""
    growth_val = parse_numeric_value(growth)
    if growth_val < 0.5:
        return f"Minimal annual rent growth of {growth_val:.1f}%"
    elif 0.5 <= growth_val < 1.25:
        return f"Moderate annual rent growth of {growth_val:.1f}%"
    elif 1.25 <= growth_val <= 2.25:
        return f"Strong annual rent growth of {growth_val:.1f}%"
    else:
        return f"Very high annual rent growth of {growth_val:.1f}%"

###############################################
# Main Function to Write to the Template      #
###############################################

def write_to_template(extracted_fields, template_path, output_path):
    """
    Writes the extracted fields into a preloaded spreadsheet template.
    
    Args:
        extracted_fields (dict): Dictionary from your extraction output.
        template_path (str): Path to the Excel template.
        output_path (str): Path where the filled-out spreadsheet will be saved.
    
    Returns:
        output_path (str): Path to the saved Excel file.
    """
    wb = load_workbook(template_path)
    ws = wb.active

    def write_comment(cell_range: str, comment: str):
        """Helper function to write to merged cells"""
        # Get the first cell of the merged range (top-left)
        first_cell = cell_range.split(':')[0]
        ws[first_cell] = comment

    # Get common values used in multiple places
    address_raw = extracted_fields.get("Address", {})
    current_tenant = extracted_fields.get("Current Tenant", "")
    restaurant_flag = extracted_fields.get("Restaurant/Auto/Medical?", "")

    # Address: cell C3
    formatted_address = map_address(address_raw, current_tenant)
    ws["C3"] = formatted_address

    # Restaurant/Auto/Medical?: cell L8 + comment in N7:N8
    ws["L8"] = map_restaurant_auto_medical(restaurant_flag)
    write_comment("N7:N8", map_restaurant_auto_medical_comment(current_tenant, restaurant_flag))

    # Single Tenant?: cell L10 + comment in N9:N10
    single_tenant_raw = extracted_fields.get("Single Tenant?", "")
    ws["L10"] = map_single_tenant(single_tenant_raw)
    write_comment("N9:N10", map_single_tenant_comment(single_tenant_raw))

    # Portfolio Management Target Brand: cell L12 + comment in N11:N12
    ws["L12"] = map_portfolio_target()
    write_comment("N11:N12", map_portfolio_target_brand_comment(current_tenant))

    # Portfolio Management Target Geography: cell L14 + comment in N13:N14
    ws["L14"] = map_portfolio_target()
    write_comment("N13:N14", map_portfolio_target_geography_comment(address_raw))

    # Acreage: cell L30 + comment in N29:N30
    acreage_raw = extracted_fields.get("Acreage", "")
    ws["L30"] = map_acreage(acreage_raw)
    write_comment("N29:N30", map_acreage_comment(acreage_raw))

    # Drive-Thru (QSR) / Carry-out (CDR): cell L36 + comment in N35:N36
    drive_val = extracted_fields.get("Drive-Thru (QSR) / Carry-out (CDR)", "") or ""
    ws["L36"] = map_drive_thru_carryout(drive_val)
    write_comment("N35:N36", map_drive_thru_comment(drive_val, current_tenant))

    # Box Size: cell L38 + comment in N37:N38
    box_size_raw = extracted_fields.get("Box Size", "")
    ws["L38"] = map_box_size(box_size_raw)
    write_comment("N37:N38", map_box_size_comment(box_size_raw))

    # Number of National Locations: cell L62 + comment in N61:N62
    concept_value = map_national_locations(extracted_fields)
    ws["L62"] = concept_value
    write_comment("N60:N62", map_national_locations_comment(
        extracted_fields.get("Number of National Locations", ""),
        current_tenant
    ))

    # Lease Structure: cell L72 + comment in N71:N72
    lease_structure_raw = extracted_fields.get("Lease Structure", "")
    ws["L72"] = map_lease_structure(lease_structure_raw)
    write_comment("N71:N72", map_lease_structure_comment(lease_structure_raw))

    # Lease Term: cell L74 + comment in N73:N74
    lease_term_raw = extracted_fields.get("Lease Term", "")
    ws["L74"] = map_lease_term(lease_term_raw)
    write_comment("N73:N74", map_lease_term_comment(lease_term_raw))

    # Absolute Rent: cell L78 + comment in N77:N78
    absolute_rent_raw = extracted_fields.get("Absolute Rent", "")
    drive_val_lower = drive_val.lower() if isinstance(drive_val, str) else ""
    building_type = "CDR" if "cdr" in drive_val_lower else "QSR"
    ws["L78"] = map_absolute_rent(absolute_rent_raw, building_type)
    write_comment("N75:N78", map_absolute_rent_comment(absolute_rent_raw, building_type))

    # Rent Growth: cell L81 + comment in N80:N81
    rent_growth_raw = extracted_fields.get("Rent Growth", "")
    ws["L81"] = map_rent_growth(rent_growth_raw)
    write_comment("N79:N81", map_rent_growth_comment(rent_growth_raw))

    wb.save(output_path)
    return output_path

def sanitize_filename(name: str) -> str:
    """
    Sanitize a string to be used as a filename by replacing invalid characters.
    """
    # Replace problematic characters with safe alternatives
    replacements = {
        '/': '-',    # Forward slash
        '\\': '-',   # Backslash
        ':': '-',    # Colon
        '*': '_',    # Asterisk
        '?': '_',    # Question mark
        '"': "'",    # Double quote
        '<': '(',    # Less than
        '>': ')',    # Greater than
        '|': '-',    # Pipe
        '\n': ' ',   # Newline
        '\r': ' ',   # Carriage return
        '\t': ' ',   # Tab
    }
    
    for char, replacement in replacements.items():
        name = name.replace(char, replacement)
    
    # Remove any other non-printable characters
    name = ''.join(char for char in name if char.isprintable())
    
    # Trim spaces and dots from the ends
    name = name.strip('. ')
    
    return name

def get_template_from_s3(bucket_name: str, template_key: str) -> io.BytesIO:
    """
    Downloads the template file from S3 into memory.
    
    Args:
        bucket_name: Name of the S3 bucket
        template_key: S3 key for the template file
        
    Returns:
        BytesIO object containing the template file
    """
    try:
        s3 = boto3.client(
            's3',
            aws_access_key_id=os.getenv('AWS_ACCESS_KEY_ID'),
            aws_secret_access_key=os.getenv('AWS_SECRET_ACCESS_KEY')
        )
        
        # Download template to memory
        template_obj = io.BytesIO()
        s3.download_fileobj(bucket_name, template_key, template_obj)
        template_obj.seek(0)  # Reset file pointer to beginning
        return template_obj
    except ClientError as e:
        print(f"Error downloading template: {e}")
        raise

def build_scorecard(
    source: str | Path,
    template_path: str | Path = None,  # Made optional
    *,
    settings_list: list[dict] | None = None,
    keywords: list[str] | None = None,
    out_dir: str | Path = "/tmp",
    client: Optional[OpenAI] = None,
) -> tuple[dict, str]:
    """
    Parse an OM PDF or plain text, extract the fields via GPT, and write a filled‑out Excel scorecard.
    
    Now supports loading template from S3.
    """
    from datetime import datetime
    from pathlib import Path
    import os
    from openai import OpenAI

    out_dir = Path(out_dir)

    if client is None:
        client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

    # Load template from S3 if path not provided
    if template_path is None:
        bucket = os.getenv('S3_BUCKET_NAME')
        key = os.getenv('TEMPLATE_S3_KEY', 'templates/Scorecard - Blank v1 streamlit.xlsx')
        template_obj = get_template_from_s3(bucket, key)
        wb = load_workbook(template_obj)
    else:
        wb = load_workbook(template_path)

    # --- defaults --------------------------------------------------
    if settings_list is None:
        settings_list = [
            {"vertical_strategy": "lines", "horizontal_strategy": "lines",
             "intersection_x_tolerance": 10, "intersection_y_tolerance": 10},
            {"vertical_strategy": "lines", "horizontal_strategy": "lines",
             "intersection_x_tolerance": 25, "intersection_y_tolerance": 25},
            {"vertical_strategy": "text",  "horizontal_strategy": "text",
             "intersection_x_tolerance": 15, "intersection_y_tolerance": 15},
        ]
    keywords = KW if keywords is None else keywords

    # --- 1) Get text payload --------------------------------------
    if isinstance(source, str):
        # Direct text input
        payload = source
        source_name = "email_text"
    elif isinstance(source, Path):
        if source.suffix.lower() == '.pdf':
            # PDF file
            print("📄 Reading PDF:", source.resolve())
            payload = get_best_payload(source, settings_list=settings_list, keywords=keywords)
            source_name = source.stem
        else:
            # Text file
            payload = source.read_text()
            source_name = source.stem
    else:
        raise TypeError(f"Expected str or Path, got {type(source)}")

    print("Here's the extracted text:", payload)

    # --- 2) LLM interpretation  -----------------------------------
    result_raw = interpret_payload_with_gpt(payload, client=client)
    result = normalize_fields(result_raw)
    print(f"Here are the extracted results: \n {result} \n")

    # --- 3) build output file name --------------------------------
    addr = result.get("Address", {}) if isinstance(result.get("Address"), dict) else {}
    tenant = result.get("Current Tenant", "Unknown Tenant")
    city_state = f"{addr.get('City','')}, {addr.get('State','')}".strip(" ,")

    ts = datetime.now().strftime("%Y.%m.%d")
    # Sanitize the tenant name and location for the filename
    safe_tenant = sanitize_filename(tenant)
    safe_location = sanitize_filename(city_state)
    
    out_name = f"Auto Scorecard - {safe_tenant} ({safe_location}) {ts} v1.xlsx"
    out_path = out_dir / out_name

    # --- 4) write to template -------------------------------------
    write_to_template(result, template_path, out_path)

    return result, str(out_path)


